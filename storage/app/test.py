import requests
import pandas as pd
import logging
import os,sys,random
from PIL import Image  # Only import Image
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

temp_image_dir = 'storage/temp_images'
os.makedirs(temp_image_dir, exist_ok=True)

def flatten_json(nested_json, exclude_keys=None):
    if exclude_keys is None:
        exclude_keys = []
    out = {}

    def flatten(x, name=''):
        if type(x) is dict:
            for a in x:
                if a not in exclude_keys:
                    flatten(x[a], name + a + '_')
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + '_')
                i += 1
        else:
            out[name[:-1]] = x

    flatten(nested_json)
    return out

def write_to_excel(data, sheet_name, file_path):
    df = pd.DataFrame([data])
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def download_image(url, save_dir):
  try:
    response = requests.get(url)
    response.raise_for_status()
    img_name = os.path.join(save_dir, os.path.basename(url))
    with open(img_name, 'wb') as img_file:
      img_file.write(response.content)
    return img_name
  except requests.RequestException as e:
    return None


def resize_image(image_path, frame_width, frame_height):
  with Image.open(image_path) as img:
    img_ratio = img.width / img.height
    frame_ratio = frame_width / frame_height

    if img_ratio > frame_ratio:
      # Image is wider than frame
      new_width = frame_width
      new_height = int(frame_width / img_ratio)
    else:
      # Image is taller than frame
      new_height = frame_height
      new_width = int(frame_height * img_ratio)

    # Use BICUBIC for resampling
    resized_img = img.resize((new_width, new_height), Image.BICUBIC)
    resized_img.save(image_path)



def create_photos_sheet_with_images(photo_data, file_path, start_cell='A1'):
    try:
        wb = load_workbook(file_path)
        photos_sheet_name = 'Photos'

        if photos_sheet_name not in wb.sheetnames:
            wb.create_sheet(photos_sheet_name)

        ws_photos = wb[photos_sheet_name]
        ws_sheet1 = wb['Sheet1']  # Accessing Sheet1

        image_info_list = []  # List to store cell numbers of images, types, and descriptions

        frame_width = 600
        frame_height = 300
        start_row = int(start_cell[1:])
        start_col = start_cell[0]

        for i, (url, description, photo_type) in enumerate(photo_data):
            img_path = download_image(url, temp_image_dir)
            if img_path:
                resize_image(img_path, frame_width, frame_height)

                image_row_height = int(frame_height / 15)
                img_cell = f'{start_col}{start_row}'
                desc_cell = f'{start_col}{start_row + image_row_height}'
                type_cell = f'{start_col}{start_row + image_row_height + 1}'

                img = OpenpyxlImage(img_path)
                ws_photos.add_image(img, img_cell)

                ws_photos[desc_cell] = description
                ws_photos[type_cell] = photo_type

                image_info_list.append((img_cell, desc_cell, type_cell))  # Record cell numbers

                start_row += image_row_height + 2

        # Write cell numbers to Sheet1
        for i, info in enumerate(image_info_list, start=1):  # Starting from row 1 in Sheet1
            ws_sheet1[f'A{i}'] = info[0]  # Photo cell number
            ws_sheet1[f'B{i}'] = info[1]  # Description cell number
            ws_sheet1[f'C{i}'] = info[2]  # Type cell number
        for sheet in wb.sheetnames:
            wb[sheet].protection.sheet = True
            wb[sheet].protection.set_password('0000')
        wb.save(file_path)
    except Exception as e:
        return


f_name = 'public/' + sys.argv[1]
file = open(f_name,'r')
data = file.read()
file.close()
random_string = generate_random_string()
excel_file_path = 'public/' + random_string + '.xlsx'
data = json.loads(data)

if not os.path.exists(excel_file_path):
    pd.DataFrame().to_excel(excel_file_path)


if data:
    # General Info
    general_info_keys = ['name', 'address', 'address_2', 'city', 'state', 'zip', 'country', 'overall_rating', 'rating_scale', 'inspection_date', 'primary_type', 'secondary_type']
    general_info = {key: data[key] for key in general_info_keys if key in data}
    write_to_excel(general_info, 'General Info', excel_file_path)
    # Physical conditions and DM
    physical_conditions_dm = flatten_json(data.get('physical_condition', {}))
    write_to_excel(physical_conditions_dm, 'Physical conditions and DM', excel_file_path)

    # Photos
    photo_data = [(photo.get('photo_url'), photo.get('photo_description'), photo.get('photo_type'))
                  for photo in data.get('images', [])
                  if all(k in photo for k in ['photo_url', 'photo_description', 'photo_type'])]

    logging.info(f"Photo Data (URLs, Descriptions, Types): {photo_data}")

    # Call the function to create the photos sheet with photo_data
    create_photos_sheet_with_images(photo_data, excel_file_path)

    # Rent Roll
    rent_roll = flatten_json(data.get('rent_roll', {}))
    write_to_excel(rent_roll, 'Rent Roll', excel_file_path)

    # Mgmt Interview
    mgmt_interview = flatten_json(data.get('mgmt_interview', {}))
    write_to_excel(mgmt_interview, 'Mgmt Interview', excel_file_path)

    # Multifamily
    multifamily = flatten_json(data.get('multifamily', {}))
    write_to_excel(multifamily, 'Multifamily', excel_file_path)
    # Fannie Mae Addendum
    fannie_mae_addendum = flatten_json(data.get('fannie_mae_assmt', {}))
    write_to_excel(fannie_mae_addendum, 'Fannie Mae Addendum', excel_file_path)

    # Fre Assmt Addendum
    fre_assmt_addendum = flatten_json(data.get('fre_assmt', {}))
    write_to_excel(fre_assmt_addendum, 'Fre Assmt Addendum', excel_file_path)

    # Repairs Verification
    repairs_verification = flatten_json(data.get('repairs_verification', {}))
    write_to_excel(repairs_verification, 'Repairs Verification', excel_file_path)

    # Senior Housing Supplement
    senior_housing_supplement = flatten_json(data.get('senior_supplement', {}))
    write_to_excel(senior_housing_supplement, 'Senior Housing Supplement', excel_file_path)

    # Hospitals
    hospitals = flatten_json(data.get('hospitals', {}))
    write_to_excel(hospitals, 'Hospitals', excel_file_path)

    print(random_string)


